from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


FIELD_HINTS = {
    "date_column": (
        "成交日期",
        "订单日期",
        "日期",
        "时间",
        "创建时间",
        "支付时间",
        "date",
        "time",
    ),
    "amount_column": (
        "成交金额",
        "销售额",
        "金额",
        "gmv",
        "revenue",
        "sales",
        "amount",
    ),
    "customer_count_column": (
        "成交客户数",
        "客户数",
        "用户数",
        "人数",
        "customer count",
        "customer_count",
        "customers",
    ),
    "product_column": (
        "产品",
        "产品类型",
        "商品",
        "品类",
        "product",
        "category",
    ),
    "region_column": (
        "区域",
        "省份",
        "城市",
        "地区",
        "region",
        "province",
        "city",
    ),
    "order_id_column": (
        "订单号",
        "订单id",
        "order_id",
        "流水号",
        "编号",
    ),
}

ORANGE = "ED7D31"
DARK_ORANGE = "C65911"
LIGHT_ORANGE = "FCE4D6"
PALE_ORANGE = "FFF2CC"
WHITE = "FFFFFF"
DARK_TEXT = "3F3F3F"
LIGHT_BORDER = Side(style="thin", color="E7E6E6")


def detect_dashboard_fields(df: pd.DataFrame) -> dict[str, str | None]:
    """Detect common business fields used by the generated Excel Dashboard."""
    columns = list(df.columns)
    result: dict[str, str | None] = {}
    for field_name, hints in FIELD_HINTS.items():
        candidates = columns
        if field_name in {"amount_column", "customer_count_column"}:
            candidates = [
                column
                for column in columns
                if pd.api.types.is_numeric_dtype(df[column])
            ]
        result[field_name] = _find_matching_column(candidates, hints)

    if result["date_column"] is None:
        for column in columns:
            if pd.api.types.is_datetime64_any_dtype(df[column]):
                result["date_column"] = column
                break
    return result


def calculate_dashboard_kpis(
    df: pd.DataFrame,
    field_config: dict[str, str | None],
) -> dict[str, Any]:
    """Calculate total KPIs and latest-month month-over-month/year-over-year growth."""
    amount = _numeric_sum(df, field_config.get("amount_column"))
    customers = _customer_count(df, field_config)
    orders = _order_count(df, field_config)
    unit_price = amount / customers if customers else None
    mom = None
    yoy = None

    dated = _with_valid_dates(df, field_config.get("date_column"))
    if not dated.empty and field_config.get("amount_column"):
        dated["_month"] = dated["_dashboard_date"].dt.to_period("M")
        periods = sorted(dated["_month"].dropna().unique())
        if periods:
            current_period = periods[-1]
            current_amount = _numeric_sum(
                dated.loc[dated["_month"] == current_period],
                field_config.get("amount_column"),
            )
            previous_amount = _numeric_sum(
                dated.loc[dated["_month"] == current_period - 1],
                field_config.get("amount_column"),
            )
            previous_year_amount = _numeric_sum(
                dated.loc[dated["_month"] == current_period - 12],
                field_config.get("amount_column"),
            )
            mom = _growth(current_amount, previous_amount)
            yoy = _growth(current_amount, previous_year_amount)

    return {
        "成交金额": amount if field_config.get("amount_column") else None,
        "成交客户数": customers,
        "客单价": unit_price,
        "订单数": orders,
        "环比": mom,
        "同比": yoy,
    }


def calculate_region_summary(
    df: pd.DataFrame,
    field_config: dict[str, str | None],
) -> pd.DataFrame:
    """Aggregate amount, customers and orders by detected region."""
    return _calculate_dimension_summary(df, field_config, "region_column", "地区")


def calculate_product_summary(
    df: pd.DataFrame,
    field_config: dict[str, str | None],
) -> pd.DataFrame:
    """Aggregate amount, customers and orders by detected product."""
    return _calculate_dimension_summary(df, field_config, "product_column", "产品")


def calculate_time_trend(
    df: pd.DataFrame,
    field_config: dict[str, str | None],
) -> pd.DataFrame:
    """Aggregate business KPIs by day or month based on available date coverage."""
    dated = _with_valid_dates(df, field_config.get("date_column"))
    if dated.empty:
        return pd.DataFrame(columns=["周期", "成交金额", "成交客户数", "客单价", "订单数"])

    date_span = (
        dated["_dashboard_date"].max() - dated["_dashboard_date"].min()
    ).days
    if date_span > 62:
        dated["_period"] = dated["_dashboard_date"].dt.to_period("M").astype(str)
    else:
        dated["_period"] = dated["_dashboard_date"].dt.strftime("%Y-%m-%d")

    rows = []
    for period, group in dated.groupby("_period", sort=True):
        amount = _numeric_sum(group, field_config.get("amount_column"))
        customers = _customer_count(group, field_config)
        rows.append(
            {
                "周期": period,
                "成交金额": amount if field_config.get("amount_column") else None,
                "成交客户数": customers,
                "客单价": amount / customers if customers else None,
                "订单数": _order_count(group, field_config),
            }
        )
    return pd.DataFrame(rows)


def create_excel_dashboard(
    df: pd.DataFrame,
    output_path: str | Path | None = None,
    field_config: dict[str, str | None] | None = None,
) -> bytes:
    """Generate a standalone Excel Dashboard without pivot tables or slicers."""
    detected = detect_dashboard_fields(df)
    config = {
        key: value if value in df.columns else None
        for key, value in {**detected, **(field_config or {})}.items()
    }
    kpis = calculate_dashboard_kpis(df, config)
    region_summary = calculate_region_summary(df, config)
    product_summary = calculate_product_summary(df, config)
    time_trend = calculate_time_trend(df, config)
    quality_summary = _calculate_quality_summary(df)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    processed = workbook.create_sheet("Processed Data")
    region_sheet = workbook.create_sheet("Region Summary")
    product_sheet = workbook.create_sheet("Product Summary")
    trend_sheet = workbook.create_sheet("Time Trend")
    quality_sheet = workbook.create_sheet("Data Quality Summary")

    _write_dataframe_sheet(processed, df)
    _write_summary_or_message(
        region_sheet,
        region_summary,
        "未检测到区域字段，无法生成区域分析。",
    )
    _write_summary_or_message(
        product_sheet,
        product_summary,
        "未检测到产品字段，无法生成产品分析。",
    )
    _write_summary_or_message(
        trend_sheet,
        time_trend,
        "未检测到日期字段，无法生成时间趋势。",
    )
    _write_dataframe_sheet(quality_sheet, quality_summary)
    _build_dashboard_sheet(
        dashboard,
        kpis,
        region_sheet,
        region_summary,
        product_sheet,
        product_summary,
        trend_sheet,
        time_trend,
        config,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    result = buffer.getvalue()
    if output_path is not None:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(result)
    return result


def _find_matching_column(columns, hints) -> str | None:
    normalized_columns = [(column, str(column).strip().casefold()) for column in columns]
    for hint in hints:
        normalized_hint = hint.casefold()
        for column, normalized_column in normalized_columns:
            if normalized_column == normalized_hint:
                return column
    for hint in hints:
        normalized_hint = hint.casefold()
        for column, normalized_column in normalized_columns:
            if normalized_hint in normalized_column:
                return column
    return None


def _numeric_sum(df: pd.DataFrame, column: str | None) -> float:
    if not column or column not in df.columns:
        return 0.0
    return float(pd.to_numeric(df[column], errors="coerce").sum())


def _customer_count(df: pd.DataFrame, config: dict[str, str | None]) -> float | None:
    column = config.get("customer_count_column")
    if not column or column not in df.columns:
        return None
    return _numeric_sum(df, column)


def _order_count(df: pd.DataFrame, config: dict[str, str | None]) -> int:
    column = config.get("order_id_column")
    if column and column in df.columns:
        return int(df[column].nunique(dropna=True))
    return int(len(df))


def _growth(current: float | None, previous: float | None) -> float | None:
    if current is None or previous is None or previous == 0:
        return None
    return (current - previous) / previous


def _with_valid_dates(df: pd.DataFrame, date_column: str | None) -> pd.DataFrame:
    if not date_column or date_column not in df.columns:
        return pd.DataFrame()
    result = df.copy()
    result["_dashboard_date"] = pd.to_datetime(result[date_column], errors="coerce")
    return result.dropna(subset=["_dashboard_date"])


def _calculate_dimension_summary(
    df: pd.DataFrame,
    config: dict[str, str | None],
    dimension_key: str,
    dimension_label: str,
) -> pd.DataFrame:
    dimension = config.get(dimension_key)
    if not dimension or dimension not in df.columns:
        return pd.DataFrame()
    rows = []
    for value, group in df.groupby(dimension, dropna=False):
        amount = _numeric_sum(group, config.get("amount_column"))
        customers = _customer_count(group, config)
        rows.append(
            {
                dimension_label: "未填写" if pd.isna(value) else value,
                "成交金额": amount if config.get("amount_column") else None,
                "成交客户数": customers,
                "客单价": amount / customers if customers else None,
                "订单数": _order_count(group, config),
            }
        )
    result = pd.DataFrame(rows)
    sort_column = "成交金额" if config.get("amount_column") else "订单数"
    return result.sort_values(sort_column, ascending=False, ignore_index=True)


def _calculate_quality_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {"指标": "数据行数", "结果": len(df)},
        {"指标": "字段数量", "结果": len(df.columns)},
        {"指标": "缺失值总数", "结果": int(df.isna().sum().sum())},
        {"指标": "重复行数量", "结果": int(df.duplicated().sum())},
    ]
    for column in df.columns:
        missing = int(df[column].isna().sum())
        if missing:
            rows.append(
                {
                    "指标": f"{column} 缺失值",
                    "结果": missing,
                }
            )
    return pd.DataFrame(rows)


def _write_summary_or_message(
    worksheet,
    dataframe: pd.DataFrame,
    message: str,
) -> None:
    if dataframe.empty:
        dataframe = pd.DataFrame({"说明": [message]})
    _write_dataframe_sheet(worksheet, dataframe)


def _write_dataframe_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    frame = dataframe.copy()
    for column in frame.columns:
        if pd.api.types.is_datetime64_any_dtype(frame[column]):
            frame[column] = frame[column].dt.tz_localize(None)
        elif frame[column].dtype == "object":
            frame[column] = frame[column].map(
                lambda value: str(value)
                if isinstance(value, (dict, list, tuple, set))
                else value
            )
    frame = frame.astype(object).where(pd.notna(frame), None)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(list(row))
    _style_data_sheet(worksheet)


def _style_data_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    if worksheet.max_row:
        worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=ORANGE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=LIGHT_BORDER)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=LIGHT_BORDER)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    for index, cells in enumerate(
        worksheet.iter_cols(min_row=1, max_row=min(worksheet.max_row, 300)),
        start=1,
    ):
        longest = max((len(str(cell.value or "")) for cell in cells), default=10)
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(longest + 3, 12),
            28,
        )


def _build_dashboard_sheet(
    worksheet,
    kpis: dict[str, Any],
    region_sheet,
    region_summary: pd.DataFrame,
    product_sheet,
    product_summary: pd.DataFrame,
    trend_sheet,
    time_trend: pd.DataFrame,
    config: dict[str, str | None],
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A4"
    for column in range(1, 19):
        worksheet.column_dimensions[get_column_letter(column)].width = 11
    for row in range(1, 48):
        worksheet.row_dimensions[row].height = 22

    worksheet.merge_cells("A1:R2")
    title = worksheet["A1"]
    title.value = "DataInsight Agent 月度销售数据监控"
    title.fill = PatternFill("solid", fgColor=DARK_ORANGE)
    title.font = Font(color=WHITE, size=22, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    card_ranges = [
        ("A4:C7", "成交金额", kpis["成交金额"], "#,##0.00"),
        ("D4:F7", "成交客户数", kpis["成交客户数"], "#,##0"),
        ("G4:I7", "客单价", kpis["客单价"], "#,##0.00"),
        ("J4:L7", "订单数", kpis["订单数"], "#,##0"),
        ("M4:O7", "环比", kpis["环比"], "0.0%"),
        ("P4:R7", "同比", kpis["同比"], "0.0%"),
    ]
    for cell_range, label, value, number_format in card_ranges:
        _write_kpi_card(worksheet, cell_range, label, value, number_format)

    _write_section_title(worksheet, "A9:I9", "区域成交明细")
    _write_section_title(worksheet, "J9:R9", "成交时间趋势")
    _write_section_title(worksheet, "A29:I29", "产品成交明细")
    _write_section_title(worksheet, "J29:R29", "Dashboard 字段设置")

    if (
        config.get("amount_column")
        and not region_summary.empty
        and "成交金额" in region_summary
    ):
        _add_bar_chart(
            worksheet,
            region_sheet,
            min(len(region_summary), 10),
            "区域成交金额 Top 10",
            "A10",
        )
    else:
        _write_dashboard_message(worksheet, "A11:I13", "未检测到区域字段或金额字段，无法生成区域分析。")

    if (
        config.get("amount_column")
        and not time_trend.empty
        and "成交金额" in time_trend
    ):
        _add_line_chart(
            worksheet,
            trend_sheet,
            len(time_trend),
            "成交金额趋势",
            "J10",
        )
    else:
        _write_dashboard_message(worksheet, "J11:R13", "未检测到日期字段或金额字段，无法生成时间趋势。")

    if (
        config.get("amount_column")
        and not product_summary.empty
        and "成交金额" in product_summary
    ):
        _add_bar_chart(
            worksheet,
            product_sheet,
            min(len(product_summary), 10),
            "产品成交金额 Top 10",
            "A30",
        )
    else:
        _write_dashboard_message(worksheet, "A31:I33", "未检测到产品字段或金额字段，无法生成产品分析。")

    field_rows = [
        ("日期字段", config.get("date_column")),
        ("金额字段", config.get("amount_column")),
        ("客户数字段", config.get("customer_count_column")),
        ("产品字段", config.get("product_column")),
        ("地区字段", config.get("region_column")),
        ("订单ID字段", config.get("order_id_column")),
    ]
    for offset, (label, value) in enumerate(field_rows, start=31):
        worksheet[f"J{offset}"] = label
        worksheet[f"J{offset}"].font = Font(bold=True, color=DARK_ORANGE)
        worksheet[f"M{offset}"] = value or "未识别"
        worksheet[f"M{offset}"].alignment = Alignment(horizontal="left")


def _write_kpi_card(
    worksheet,
    cell_range: str,
    label: str,
    value: Any,
    number_format: str,
) -> None:
    worksheet.merge_cells(cell_range)
    start_cell = worksheet[cell_range.split(":")[0]]
    display_value = value
    if value is None:
        display_value = "暂无数据"
    start_cell.value = f"{label}\n{display_value}" if isinstance(display_value, str) else display_value
    start_cell.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    start_cell.font = Font(color=DARK_TEXT, size=16, bold=True)
    start_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    start_cell.border = Border(
        left=LIGHT_BORDER,
        right=LIGHT_BORDER,
        top=LIGHT_BORDER,
        bottom=LIGHT_BORDER,
    )
    if not isinstance(display_value, str):
        start_cell.number_format = number_format
        start_cell.value = f"{label}\n{_format_kpi_value(display_value, number_format)}"


def _format_kpi_value(value: float | int, number_format: str) -> str:
    if "%" in number_format:
        return f"{value:.1%}"
    if "." in number_format:
        return f"{value:,.2f}"
    return f"{value:,.0f}"


def _write_section_title(worksheet, cell_range: str, text: str) -> None:
    worksheet.merge_cells(cell_range)
    cell = worksheet[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=ORANGE)
    cell.font = Font(color=WHITE, size=13, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_dashboard_message(worksheet, cell_range: str, message: str) -> None:
    worksheet.merge_cells(cell_range)
    cell = worksheet[cell_range.split(":")[0]]
    cell.value = message
    cell.fill = PatternFill("solid", fgColor=PALE_ORANGE)
    cell.font = Font(color=DARK_TEXT, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_bar_chart(
    dashboard,
    source_sheet,
    row_count: int,
    title: str,
    anchor: str,
) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.height = 9
    chart.width = 17
    chart.legend = None
    data = Reference(source_sheet, min_col=2, min_row=1, max_row=row_count + 1)
    categories = Reference(source_sheet, min_col=1, min_row=2, max_row=row_count + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    dashboard.add_chart(chart, anchor)


def _add_line_chart(
    dashboard,
    source_sheet,
    row_count: int,
    title: str,
    anchor: str,
) -> None:
    chart = LineChart()
    chart.style = 13
    chart.title = title
    chart.height = 9
    chart.width = 17
    chart.y_axis.title = "成交金额"
    data = Reference(source_sheet, min_col=2, min_row=1, max_row=row_count + 1)
    categories = Reference(source_sheet, min_col=1, min_row=2, max_row=row_count + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    dashboard.add_chart(chart, anchor)
