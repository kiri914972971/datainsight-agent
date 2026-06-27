from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_TEMPLATE_PATH = PROJECT_ROOT / "templates" / "default_dashboard.xlsx"
UPLOADED_TEMPLATE_STATE_KEY = "uploaded_report_template"

TEMPLATE_TYPE_LABELS = {
    "excel_dashboard_template": "Excel 仪表盘模板",
    "word_template": "Word 模板",
    "ppt_template": "PPT 模板",
    "unsupported": "不支持的模板",
}


def get_last_sheet_name(workbook) -> str:
    """Return the final worksheet name without assuming a fixed dashboard name."""
    if not workbook.sheetnames:
        raise ValueError("Excel 模板不包含任何工作表。")
    return workbook.sheetnames[-1]


def _template_info(
    template_type: str,
    file_name: str,
    source: str,
    sheet_name: str | None = None,
) -> dict:
    return {
        "template_type": template_type,
        "template_type_label": TEMPLATE_TYPE_LABELS[template_type],
        "file_name": file_name,
        "sheet_name": sheet_name,
        "source": source,
    }


def load_default_excel_template():
    """Load the built-in Excel template and select its final worksheet."""
    if not DEFAULT_EXCEL_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            "未找到默认模板，请将文件放到 templates/default_dashboard.xlsx。"
        )

    workbook = load_workbook(
        DEFAULT_EXCEL_TEMPLATE_PATH,
        read_only=True,
        data_only=False,
    )
    sheet_name = get_last_sheet_name(workbook)
    template_info = _template_info(
        "excel_dashboard_template",
        DEFAULT_EXCEL_TEMPLATE_PATH.name,
        "default",
        sheet_name,
    )
    template_info["path"] = "templates/default_dashboard.xlsx"
    return workbook, sheet_name, template_info


def detect_template_type(uploaded_file) -> str:
    """Identify a report template from its file extension."""
    file_name = getattr(uploaded_file, "name", str(uploaded_file))
    suffix = Path(file_name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return "excel_dashboard_template"
    if suffix == ".docx":
        return "word_template"
    if suffix == ".pptx":
        return "ppt_template"
    return "unsupported"


def _read_uploaded_bytes(uploaded_file) -> bytes:
    if hasattr(uploaded_file, "getvalue"):
        return uploaded_file.getvalue()
    if hasattr(uploaded_file, "read"):
        return uploaded_file.read()
    raise TypeError("无法读取上传的模板文件。")


def _load_excel_bytes(file_bytes: bytes, file_name: str):
    return load_workbook(
        BytesIO(file_bytes),
        read_only=True,
        data_only=False,
        keep_vba=Path(file_name).suffix.lower() == ".xlsm",
    )


def save_uploaded_template(uploaded_file) -> dict:
    """Save an uploaded template only in the current Streamlit session."""
    template_type = detect_template_type(uploaded_file)
    if template_type == "unsupported":
        raise ValueError("仅支持 .xlsx、.xlsm、.docx 和 .pptx 模板。")

    file_name = uploaded_file.name
    file_bytes = _read_uploaded_bytes(uploaded_file)
    sheet_name = None
    if template_type == "excel_dashboard_template":
        workbook = _load_excel_bytes(file_bytes, file_name)
        sheet_name = get_last_sheet_name(workbook)
        workbook.close()

    template_info = _template_info(
        template_type,
        file_name,
        "uploaded",
        sheet_name,
    )
    st.session_state[UPLOADED_TEMPLATE_STATE_KEY] = {
        "file_bytes": file_bytes,
        "template_info": template_info,
    }
    return template_info


def get_active_template() -> dict:
    """Return the user template when selected, otherwise the built-in template."""
    mode = st.session_state.get("report_template_mode", "使用默认模板")
    uploaded_template = st.session_state.get(UPLOADED_TEMPLATE_STATE_KEY)
    if mode == "上传自定义模板" and uploaded_template:
        template_info = uploaded_template["template_info"]
        workbook = None
        if template_info["template_type"] == "excel_dashboard_template":
            workbook = _load_excel_bytes(
                uploaded_template["file_bytes"],
                template_info["file_name"],
            )
        return {
            "workbook": workbook,
            "sheet_name": template_info.get("sheet_name"),
            "template_info": template_info,
            "file_bytes": uploaded_template["file_bytes"],
        }

    workbook, sheet_name, template_info = load_default_excel_template()
    return {
        "workbook": workbook,
        "sheet_name": sheet_name,
        "template_info": template_info,
        "file_bytes": DEFAULT_EXCEL_TEMPLATE_PATH.read_bytes(),
    }
